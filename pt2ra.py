import openpyxl
import os
import warnings
from colorama import init, Fore

# Initialize colorama
init(autoreset=True)

warnings.filterwarnings("ignore", category=UserWarning, module="openpyxl")


def process_pt_results(file_path='pt_results.xlsx', sheet_name='Risk Register'):
    """
    Process PT (Penetration Testing) results from an Excel file.

    This function reads data from a specified Excel file and sheet, extracting
    information from columns 'S/N', 'Overall Risk Rating', 'Observations', and 'Implications'.
    It processes rows until a blank row is encountered.

    Args:
        file_path (str): Path to the Excel file. Defaults to 'pt_results.xlsx'.
        sheet_name (str): Name of the sheet to process. Defaults to 'Risk Register'.

    Returns:
        list of dict or None: A list of dictionaries, where each dictionary represents a row
        with keys 'S/N', 'Overall Risk Rating', 'Observations', and 'Implications'.
        Returns None if no data is found or an error occurs.
    """
    try:
        # Check if the file exists
        if not os.path.exists(file_path):
            raise FileNotFoundError(f"The file '{file_path}' does not exist in the current directory.")

        # Load the workbook and select the sheet
        workbook = openpyxl.load_workbook(file_path)
        sheet = workbook[sheet_name]

        # Ask user to input the correct row number for headers
        header_row = 2

        # Get the header row
        headers = [cell.value for cell in sheet[header_row]]
        #print("\nHeaders found:", ", ".join(headers))

        # Define the required columns
        required_columns = ['S/N', 'Overall Risk Rating', 'Issue Title', 'Observation', 'Implications']

        column_mapping = {
            'S/N': 'S/N',
            'Overall Risk Rating': 'Overall Risk Rating',
            'Issue Title': 'Issue Title',
            'Observation': 'Observation',
            'Implications': 'Implications'
        }
        # Find column indices
        column_indices = {req_col: headers.index(actual_col) for req_col, actual_col in column_mapping.items() if actual_col in headers}

        # Initialize an empty list to store the results
        result = []

        # Loop through rows until a blank row is encountered
        for row in sheet.iter_rows(min_row=header_row+1, values_only=True):
            if all(cell is None or cell == '' for cell in row):
                print(f"Blank row encountered. Stopping the loop.")
                break
            
            row_data = {req_col: row[col_index] for req_col, col_index in column_indices.items()}
            result.append(row_data)

        if not result:
            print(Fore.LIGHTRED_EX + "No data found after processing.")
            return None

        return result  # Return the result instead of printing it

    except FileNotFoundError as e:
        print(Fore.LIGHTRED_EX + f"Error: {e}")
    except ValueError as e:
        print(Fore.LIGHTRED_EX + f"Invalid input: {e}")
    except Exception as e:
        print(Fore.LIGHTRED_EX + f"An unexpected error occurred: {e}")
    
    return None  # Return None if an exception occurs

def fill_ra_template(processed_data, template_path='RA_Blank_Template_Only.xlsx', template_sheet='Risk Assessment Template'):
    """
    Fill the Risk Assessment template with processed data.

    Args:
        processed_data (list of dict): Processed data from PT results.
        template_path (str): Path to the RA template Excel file.
        template_sheet (str): Name of the sheet in the RA template.
    """
    try:
        # Load the RA template workbook
        workbook = openpyxl.load_workbook(template_path)
        sheet = workbook[template_sheet]

        # Check if the template structure is correct
        expected_values = {
            'B5': 'Title',
            'C5': 'Risk Statement (1)',
            'G6': 'Risk Rating (4)'
        }
        
        missing_values = []
        for cell, expected_value in expected_values.items():
            if sheet[cell].value != expected_value:
                missing_values.append(f"{cell}: Expected '{expected_value}', found '{sheet[cell].value}'")
        
        if missing_values:
            print(Fore.LIGHTRED_EX + "The template structure is not as expected. Missing or incorrect values:")
            for missing in missing_values:
                print(Fore.LIGHTRED_EX + missing)
            raise ValueError("Please check the template structure.")

        # Start filling data from row 7
        print(Fore.LIGHTBLUE_EX + "Filling the RA template with the processed data...")
        for row, item in enumerate(processed_data, start=7):
            sheet.cell(row=row, column=2, value=item['Issue Title'])  # Column B
            sheet.cell(row=row, column=3, value=item['Implications'])  # Column C
            sheet.cell(row=row, column=7, value=item['Overall Risk Rating'])  # Column G

        # Save the filled template
        filled_template_path = 'Filled_RA_Template.xlsx'
        workbook.save(filled_template_path)
        print(Fore.LIGHTBLUE_EX + f"Filled template saved as {filled_template_path}")

    except ValueError as e:
        print(Fore.LIGHTRED_EX + f"Error: {e}")
    except Exception as e:
        print(Fore.LIGHTRED_EX + f"An error occurred while filling the RA template: {e}")

def find_cell(wb, search_string, max_rows=4):
    """
    Find the cell containing a specific string within the first max_rows rows of each worksheet.

    Args:
        wb (Workbook): The workbook to search in.
        search_string (str): The string to search for.
        max_rows (int, optional): The maximum number of rows to search. Defaults to 4.

    Returns:
        tuple or None: Returns (row, column) if found, otherwise None.
    """
    for sheet in wb.worksheets:
        for row in range(1, max_rows + 1):
            for cell in sheet[row]:
                if cell.value == search_string:
                    return row, f"{cell.column_letter}{cell.row}"
    return None

def find_cell_xy(wb, search_string, max_rows=4):
    """
    Find the cell containing a specific string within the first max_rows rows of each worksheet.

    Args:
        wb (Workbook): The workbook to search in.
        search_string (str): The string to search for.
        max_rows (int, optional): The maximum number of rows to search. Defaults to 4.

    Returns:
        tuple or None: Returns (row, column) if found, otherwise None.
    """
    for sheet in wb.worksheets:
        for row in range(1, max_rows + 1):
            for col in range(1, sheet.max_column + 1):
                cell = sheet.cell(row=row, column=col)
                if cell.value == search_string:
                    return row, col
    return None

def print_formatted_data(file_path, sheet_name):
    """
    Print formatted data from a specified sheet in an Excel file.

    This function reads data from the given Excel file and sheet, then prints
    each row's key-value pairs in a formatted, color-coded manner. It uses
    cyan separators between items, yellow for keys, and white for values.

    Args:
        file_path (str): The path to the Excel file.
        sheet_name (str): The name of the sheet to read data from.

    Returns:
        None
    """
    # Replace the undefined function with a call to process_pt_results
    data = process_pt_results(file_path, sheet_name)
    
    for item in data:
        print(Fore.CYAN + "=" * 50)
        for key, value in item.items():
            print(f"{Fore.YELLOW}{key}: {Fore.WHITE}{value}")

if __name__ == "__main__":
    def main():
        file_path = "pt_results.xlsx"  # Specify your file path
        sheet_name = "Risk Register"  # Specify your sheet name
        print(Fore.LIGHTBLUE_EX + "\nHere are the PT results:")
        data = process_pt_results(file_path, sheet_name)
        if data:  # Check if data is not None
            print_formatted_data(file_path, sheet_name)
            # Fill the RA template with processed data
            fill_ra_template(data)

    main()  # Call the main functions

